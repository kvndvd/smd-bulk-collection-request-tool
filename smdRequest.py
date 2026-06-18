from __future__ import annotations

import csv
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


SOURCE_HEADERS = {
    "CaseKey",
    "EntryDate",
    "addEntryNumber",
    "addDocURL",
    "EntryText",
}

FIRST_DUP_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
SECOND_DUP_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
FILTERED_ROW_FONT = Font(color="FF0000")

COUNSEL_FOLDER = "Counsel Request"
COURT_FOLDER = "Court Request"

DEFAULT_USER_TYPE = "SMD"
DEFAULT_HOST_SYSTEM = "FED_APP"
DEFAULT_CASE_TYPE = ""

# Add more phrases here any time in the future.
EXCLUDED_ENTRYTEXT_PHRASES = [
    "motion to extend the time",
    "motion for extension of time",
]


@dataclass
class BuildResult:
    cnl_rows: list[dict[str, str]]
    ct_rows: list[dict[str, str]]
    ct_example_rows: list[dict[str, str]]
    first_duplicate_rows: set[int] = field(default_factory=set)
    later_duplicate_rows: set[int] = field(default_factory=set)
    filtered_rows: set[int] = field(default_factory=set)


@dataclass
class OutputPaths:
    cnl_path: Path
    ct_path: Path
    xlsm_path: Path
    ct_example_path: Path | None = None


def format_filename_date(base_date: date) -> str:
    return f"{base_date:%Y-%m-%d}"


def format_example_filename_date(base_date: date) -> str:
    return f"{base_date:%Y-%m-%d}"


def format_xlsm_filename(base_date: date) -> str:
    return f"{base_date:%m%d%Y}.xlsm"


def normalize_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def normalized_dedupe_key(value: str) -> str:
    return " ".join(value.strip().split())


def parse_excel_date(value) -> date | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    formats = [
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y-%m-%d",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y %H:%M:%S",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def entry_text_matches_exclusion(entry_text: str) -> bool:
    lowered = entry_text.casefold()
    return any(phrase.casefold() in lowered for phrase in EXCLUDED_ENTRYTEXT_PHRASES)


def sort_rows_with_blank_proceeding_code_last(
    rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    return sorted(
        rows,
        key=lambda row: (not str(row.get("ProceedingCode", "")).strip(),),
    )


def find_header_map(worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_index, cell in enumerate(worksheet[1], start=1):
        if cell.value is not None:
            headers[str(cell.value).strip()] = col_index

    missing = SOURCE_HEADERS - set(headers)
    if missing:
        missing_text = ", ".join(sorted(missing))
        raise ValueError(f"Missing required columns: {missing_text}")
    return headers


def get_base_filename_date(template_path: Path) -> date:
    workbook = load_workbook(template_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_map = find_header_map(sheet)

    try:
        for row_index in range(2, sheet.max_row + 1):
            raw_entry_date = sheet.cell(row_index, header_map["EntryDate"]).value
            parsed_date = parse_excel_date(raw_entry_date)
            if parsed_date is not None:
                return parsed_date + timedelta(days=1)
    finally:
        workbook.close()

    return date.today() + timedelta(days=1)


def build_output_rows(template_path: Path) -> BuildResult:
    workbook = load_workbook(template_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_map = find_header_map(sheet)

    cnl_rows: list[dict[str, str]] = []
    ct_rows: list[dict[str, str]] = []
    ct_example_rows: list[dict[str, str]] = []
    seen_entry_texts: dict[str, int] = {}
    seen_doc_urls: dict[str, int] = {}
    first_duplicate_rows: set[int] = set()
    later_duplicate_rows: set[int] = set()
    filtered_rows: set[int] = set()

    for row_index in range(2, sheet.max_row + 1):
        case_key = normalize_value(sheet.cell(row_index, header_map["CaseKey"]).value).strip()
        if not case_key:
            continue

        entry_date = normalize_value(sheet.cell(row_index, header_map["EntryDate"]).value).strip()
        add_entry_number = normalize_value(sheet.cell(row_index, header_map["addEntryNumber"]).value).strip()
        add_doc_url = normalize_value(sheet.cell(row_index, header_map["addDocURL"]).value).strip()
        entry_text = normalize_value(sheet.cell(row_index, header_map["EntryText"]).value).strip()

        if entry_text_matches_exclusion(entry_text):
            filtered_rows.add(row_index)
            continue

        cnl_rows.append({"CaseKey": case_key})

        entry_text_key = normalized_dedupe_key(entry_text)
        doc_url_key = normalized_dedupe_key(add_doc_url)

        duplicate_source_rows: set[int] = set()
        if entry_text_key and entry_text_key in seen_entry_texts:
            duplicate_source_rows.add(seen_entry_texts[entry_text_key])
        if doc_url_key and doc_url_key in seen_doc_urls:
            duplicate_source_rows.add(seen_doc_urls[doc_url_key])

        if duplicate_source_rows:
            first_duplicate_rows.update(duplicate_source_rows)
            later_duplicate_rows.add(row_index)
            continue

        if entry_text_key:
            seen_entry_texts[entry_text_key] = row_index
        if doc_url_key:
            seen_doc_urls[doc_url_key] = row_index

        base_ct_row = {
            "CaseKey": case_key,
            "ProceedingDate": entry_date,
            "ProceedingCode": add_entry_number,
            "SubProceedingCode": "0",
            "DocURL": add_doc_url,
        }

        ct_rows.append(base_ct_row)
        ct_example_rows.append(
            {
                **base_ct_row,
                "UserType": DEFAULT_USER_TYPE,
                "HostSystem": DEFAULT_HOST_SYSTEM,
                "CaseType": DEFAULT_CASE_TYPE,
            }
        )

    workbook.close()
    return BuildResult(
        cnl_rows=cnl_rows,
        ct_rows=ct_rows,
        ct_example_rows=ct_example_rows,
        first_duplicate_rows=first_duplicate_rows,
        later_duplicate_rows=later_duplicate_rows,
        filtered_rows=filtered_rows,
    )


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="cp1252", errors="replace") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def highlight_rows(
    template_path: Path,
    output_path: Path,
    first_rows: set[int],
    later_rows: set[int],
    filtered_rows: set[int],
) -> None:
    workbook = load_workbook(template_path, keep_vba=True)
    sheet = workbook[workbook.sheetnames[0]]

    for row_index in sorted(first_rows):
        for col_index in range(1, sheet.max_column + 1):
            sheet.cell(row_index, col_index).fill = FIRST_DUP_FILL

    for row_index in sorted(later_rows):
        for col_index in range(1, sheet.max_column + 1):
            sheet.cell(row_index, col_index).fill = SECOND_DUP_FILL

    for row_index in sorted(filtered_rows):
        for col_index in range(1, sheet.max_column + 1):
            sheet.cell(row_index, col_index).font = FILTERED_ROW_FONT

    workbook.save(output_path)
    workbook.close()


def generate_outputs(
    template_path: Path,
    output_dir: Path,
    create_example_ct: bool = False,
) -> OutputPaths:
    result = build_output_rows(template_path)
    base_date = get_base_filename_date(template_path)
    file_date = format_filename_date(base_date)
    example_file_date = format_example_filename_date(base_date)

    counsel_dir = output_dir / COUNSEL_FOLDER
    court_dir = output_dir / COURT_FOLDER
    counsel_dir.mkdir(parents=True, exist_ok=True)
    court_dir.mkdir(parents=True, exist_ok=True)

    cnl_path = counsel_dir / f"smd_SMD_{file_date}_CNL_Request.csv"
    ct_path = court_dir / f"smd_SMD_{file_date}_CT_Request.csv"
    xlsm_path = output_dir / format_xlsm_filename(base_date)
    ct_example_path: Path | None = None

    sorted_ct_rows = sort_rows_with_blank_proceeding_code_last(result.ct_rows)
    sorted_ct_example_rows = sort_rows_with_blank_proceeding_code_last(result.ct_example_rows)

    write_csv(cnl_path, ["CaseKey"], result.cnl_rows)
    write_csv(
        ct_path,
        ["CaseKey", "ProceedingDate", "ProceedingCode", "SubProceedingCode", "DocURL"],
        sorted_ct_rows,
    )

    if create_example_ct:
        ct_example_path = court_dir / f"smd_SMD_{example_file_date}_CT_Request_1.csv"
        write_csv(
            ct_example_path,
            [
                "CaseKey",
                "ProceedingDate",
                "ProceedingCode",
                "SubProceedingCode",
                "DocURL",
                "UserType",
                "HostSystem",
                "CaseType",
            ],
            sorted_ct_example_rows,
        )

    highlight_rows(
        template_path,
        xlsm_path,
        result.first_duplicate_rows,
        result.later_duplicate_rows,
        result.filtered_rows,
    )

    return OutputPaths(
        cnl_path=cnl_path,
        ct_path=ct_path,
        xlsm_path=xlsm_path,
        ct_example_path=ct_example_path,
    )